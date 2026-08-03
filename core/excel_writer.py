import os
from datetime import date, datetime
from pathlib import Path
from typing import List
import openpyxl
from openpyxl.utils.datetime import from_excel

from core.task_parser import ParsedTask


def safe_excel_text(value: str) -> str:
    if value and value[0] in "=+-@":
        return "'" + value
    return value


def _as_date(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        converted = from_excel(value)
        return converted.date() if isinstance(converted, datetime) else converted
    return None


class ExcelWriter:
    def __init__(self, template_path: str):
        if not os.path.exists(template_path):
            raise FileNotFoundError(f"Excel 模板不存在: {template_path}")
        self.template_path = template_path
        self._wb = openpyxl.load_workbook(template_path)

    def add_task(self, sheet_name: str, task: ParsedTask, analysis: str = "") -> int:
        if sheet_name not in self._wb.sheetnames:
            raise KeyError(sheet_name)

        ws = self._wb[sheet_name]
        new_row = next(
            (
                row
                for row in range(2, ws.max_row + 1)
                if _as_date(ws.cell(row=row, column=1).value) == task.date
            ),
            ws.max_row + 1,
        )
        cell = ws.cell(row=new_row, column=1)
        cell.value = task.date
        cell.number_format = 'yyyy/m/d'
        task_text = "\n".join(f"{i+1}、{t}" for i, t in enumerate(task.tasks))
        ws.cell(row=new_row, column=2).value = safe_excel_text(task_text)
        if analysis:
            ws.cell(row=new_row, column=3).value = safe_excel_text(analysis)
        return new_row

    def save(self, output_path: str):
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        self._wb.save(output_path)

    def close(self):
        self._wb.close()

    def get_sheet_names(self) -> List[str]:
        return self._wb.sheetnames
