import os
import tempfile
from datetime import date

import openpyxl
import pytest

from core.excel_writer import ExcelWriter, safe_excel_text
from core.task_parser import ParsedTask


@pytest.fixture
def temp_template():
    """Create a temporary Excel template matching the real structure."""
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "张三"
    ws["A1"] = "安排时间"
    ws["B1"] = "任务安排"
    ws["C1"] = "情况分析"
    # Pre-fill some existing data
    ws["A2"] = 46144  # 2026-05-02 serial
    ws["B2"] = "旧任务"
    ws["C2"] = "已完成"
    wb.save(path)
    wb.close()
    return path


@pytest.fixture
def parsed_task():
    return ParsedTask(
        msg_id=1,
        date=date(2026, 5, 2),
        date_excel_serial=46144,
        raw_text="🚩5.2 任务\n1⃣ 订正作文\n2⃣ 完形填空",
        tasks=["订正作文", "完形填空，首字母填空，语法填空各一篇"],
    )


class TestExcelWriter:
    def test_write_to_existing_date(self, temp_template, parsed_task):
        output_path = temp_template.replace(".xlsx", "_out.xlsx")
        writer = ExcelWriter(temp_template)
        writer.add_task("张三", parsed_task)
        writer.save(output_path)

        wb = openpyxl.load_workbook(output_path)
        ws = wb["张三"]
        assert ws["A2"].value.date() == date(2026, 5, 2)
        assert "订正作文" in str(ws["B2"].value)
        assert "完形填空" in str(ws["B2"].value)
        assert ws["C2"].value == "已完成"  # C column preserved
        wb.close()
        os.unlink(output_path)

    def test_write_to_new_date(self, temp_template):
        new_task = ParsedTask(
            msg_id=2,
            date=date(2026, 4, 30),
            date_excel_serial=46143,
            raw_text="🚩4.30 任务\n1⃣ 语法填空",
            tasks=["语法填空"],
        )
        output_path = temp_template.replace(".xlsx", "_out2.xlsx")
        writer = ExcelWriter(temp_template)
        writer.add_task("张三", new_task)
        writer.save(output_path)

        wb = openpyxl.load_workbook(output_path)
        ws = wb["张三"]
        # Should have appended to last row
        assert ws["A3"].value.date() == date(2026, 4, 30)
        assert "语法填空" in str(ws["B3"].value)
        wb.close()
        os.unlink(output_path)

    def test_preserves_other_sheets(self, temp_template, parsed_task):
        wb = openpyxl.load_workbook(temp_template)
        ws2 = wb.create_sheet("李四")
        ws2["A1"] = "安排时间"
        ws2["B1"] = "任务安排"
        ws2["C1"] = "情况分析"
        wb.save(temp_template)
        wb.close()

        output_path = temp_template.replace(".xlsx", "_out3.xlsx")
        writer = ExcelWriter(temp_template)
        writer.add_task("张三", parsed_task)
        writer.save(output_path)

        wb = openpyxl.load_workbook(output_path)
        assert "李四" in wb.sheetnames
        wb.close()
        os.unlink(output_path)

    def test_write_to_existing_date_updates_that_row(self, temp_template, parsed_task):
        writer = ExcelWriter(temp_template)

        row = writer.add_task("张三", parsed_task, '=HYPERLINK("bad")')

        assert row == 2
        assert writer._wb["张三"]["B2"].value.startswith("1、订正作文")
        assert writer._wb["张三"]["C2"].value.startswith("'")
        writer.close()

    @pytest.mark.parametrize("value", ["=SUM(A1:A2)", "+123", "-123", "@name"])
    def test_safe_excel_text_prefixes_formula_characters(self, value):
        assert safe_excel_text(value) == "'" + value

    def test_missing_sheet_is_an_error(self, temp_template, parsed_task):
        writer = ExcelWriter(temp_template)
        with pytest.raises(KeyError, match="不存在的Sheet"):
            writer.add_task("不存在的Sheet", parsed_task)
        writer.close()

    def test_save_accepts_filename_without_parent(
        self, temp_template, parsed_task, tmp_path, monkeypatch
    ):
        monkeypatch.chdir(tmp_path)
        writer = ExcelWriter(temp_template)
        writer.add_task("张三", parsed_task)
        writer.save("output.xlsx")
        writer.close()

        assert (tmp_path / "output.xlsx").exists()

    def test_existing_date_has_one_row_and_escaped_formula_after_reopen(
        self, temp_template, parsed_task, tmp_path
    ):
        output_path = tmp_path / "persisted.xlsx"
        writer = ExcelWriter(temp_template)

        first_row = writer.add_task("张三", parsed_task, '=HYPERLINK("bad")')
        second_row = writer.add_task("张三", parsed_task, '=HYPERLINK("bad")')
        writer.save(str(output_path))
        writer.close()

        reopened = openpyxl.load_workbook(output_path, data_only=False)
        sheet = reopened["张三"]
        matching_rows = [
            row
            for row in range(2, sheet.max_row + 1)
            if sheet.cell(row=row, column=1).value.date() == parsed_task.date
        ]
        assert first_row == second_row == 2
        assert matching_rows == [2]
        assert sheet.max_row == 2
        assert sheet["C2"].value == "'=HYPERLINK(\"bad\")"
        assert sheet["C2"].data_type == "s"
        reopened.close()
