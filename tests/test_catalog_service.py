from pathlib import Path

from openpyxl import Workbook

import app as app_module
from schemas.catalog import CatalogDependencies
from schemas.wizard import WizardState
from services.catalog import filter_catalog, load_catalog


class FakeDatabase:
    def __init__(self):
        self.query_count = 0


def test_filter_catalog_reuses_loaded_catalog_without_requerying_database(tmp_path: Path):
    template_path = tmp_path / "template.xlsx"
    workbook = Workbook()
    workbook.active.title = "项目群"
    workbook.save(template_path)
    workbook.close()

    database = FakeDatabase()

    def query_chatrooms(queried_database):
        assert queried_database is database
        queried_database.query_count += 1
        return [("room-1", "项目群"), ("room-2", "运营群")]

    state = {
        "database": database,
        "wizard": WizardState(),
        "catalog_dependencies": CatalogDependencies(
            query_chatrooms=query_chatrooms,
            database_key="database",
        ),
    }

    first = load_catalog(state, str(template_path))
    second = filter_catalog(state, "项目")

    assert database.query_count == 1
    assert second == [first[0]]
    assert first[0].suggested_sheet == "项目群"


def test_catalog_records_are_cached_as_an_immutable_tuple(tmp_path: Path):
    template_path = tmp_path / "template.xlsx"
    workbook = Workbook()
    workbook.active.title = "项目群"
    workbook.save(template_path)
    workbook.close()

    state = {
        "database": FakeDatabase(),
        "catalog_dependencies": CatalogDependencies(
            query_chatrooms=lambda _: [("room-1", "项目群")],
            database_key="database",
        ),
    }

    result = load_catalog(state, str(template_path))

    assert isinstance(state["chatroom_catalog"], tuple)
    assert result is state["chatroom_catalog"]


def test_catalog_uses_injected_manual_group_sheet_map(tmp_path: Path):
    template_path = tmp_path / "template.xlsx"
    workbook = Workbook()
    workbook.active.title = "配置工作表"
    workbook.save(template_path)
    workbook.close()

    state = {
        "database": FakeDatabase(),
        "catalog_dependencies": CatalogDependencies(
            query_chatrooms=lambda _: [("room-1", "仅配置匹配群")],
            database_key="database",
            manual_group_sheet_map={"仅配置匹配群": "配置工作表"},
        ),
    }

    catalog = load_catalog(state, str(template_path))

    assert catalog[0].suggested_sheet == "配置工作表"


def test_new_production_session_binds_configured_manual_group_sheet_map(
    monkeypatch, tmp_path: Path
):
    template_path = tmp_path / "template.xlsx"
    workbook = Workbook()
    workbook.active.title = "配置工作表"
    workbook.save(template_path)
    workbook.close()
    monkeypatch.setattr(
        app_module.config.matching,
        "group_sheet_map",
        {"仅配置匹配群": "配置工作表"},
    )
    monkeypatch.setattr(
        app_module,
        "_get_chatrooms",
        lambda _: [("room-1", "仅配置匹配群")],
    )

    state = app_module._new_session_state()
    catalog = load_catalog(state, str(template_path))

    assert catalog[0].suggested_sheet == "配置工作表"


def test_new_production_session_binds_composition_root_factories(monkeypatch):
    class FakeWriter:
        def __init__(self, path):
            self.path = path

    class FakeParser:
        pass

    monkeypatch.setattr(app_module, "ExcelWriter", FakeWriter)
    monkeypatch.setattr(app_module, "TaskParser", FakeParser)

    state = app_module._new_session_state()

    catalog_dependencies = state["catalog_dependencies"]
    preview_dependencies = state["preview_dependencies"]
    assert isinstance(catalog_dependencies.excel_writer_factory("ignored"), FakeWriter)
    assert isinstance(preview_dependencies.task_parser_factory(), FakeParser)
